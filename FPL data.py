import requests
import openpyxl
from math import exp, factorial
from datetime import datetime
from openpyxl.styles import Font

# API URLs
url = "https://fantasy.premierleague.com/api/bootstrap-static/"
fixtures_url = "https://fantasy.premierleague.com/api/fixtures/"

# Fetch data from the APIs
response = requests.get(url)
data = response.json()
fixtures_response = requests.get(fixtures_url)
fixtures_data = fixtures_response.json()

# Extract the players' data
players = data['elements']
teams = {team['id']: team['name'] for team in data['teams']}

# Poisson distribution function
def poisson_prob(l, k):
    return (l**k * exp(-l)) / factorial(k)

# Calculate pFDR and fFDR for a team
def calculate_fdr(team_id, fixtures):
    current_date = datetime.now()
    team_fixtures = [
        fixture for fixture in fixtures
        if fixture['team_h'] == team_id or fixture['team_a'] == team_id
    ]
    past_fixtures = [
        fixture for fixture in team_fixtures
        if datetime.strptime(fixture['kickoff_time'], '%Y-%m-%dT%H:%M:%SZ') < current_date
    ]
    future_fixtures = [
        fixture for fixture in team_fixtures
        if datetime.strptime(fixture['kickoff_time'], '%Y-%m-%dT%H:%M:%SZ') >= current_date
    ]
    past_fdrs = [
        fixture['team_h_difficulty'] if fixture['team_h'] == team_id else fixture['team_a_difficulty']
        for fixture in sorted(past_fixtures, key=lambda x: x['kickoff_time'], reverse=True)[:4]
    ]
    future_fdrs = [
        fixture['team_h_difficulty'] if fixture['team_h'] == team_id else fixture['team_a_difficulty']
        for fixture in sorted(future_fixtures, key=lambda x: x['kickoff_time'])[:4]
    ]
    pFDR = sum(past_fdrs) / len(past_fdrs) if past_fdrs else None
    fFDR = sum(future_fdrs) / len(future_fdrs) if future_fdrs else None
    return pFDR, fFDR

# Load the existing Excel file or create a new one
try:
    workbook = openpyxl.load_workbook("players_data.xlsx")
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Remove the default sheet

# Create a dictionary to store categorized results
player_data = {
    "Attackers": [],
    "Midfielders": [],
    "Defenders": [],
    "Goalkeepers": []
}

# Loop through each player and get their data
for player in players:
    full_name = f"{player['first_name']} {player['second_name']}"
    player_id = player['id']
    team_id = player['team']
    position = player['element_type']  # Position: 1 = Goalkeeper, 2 = Defender, 3 = Midfielder, 4 = Attacker
    
    # Fetch the player's history (last 4 games) from the API
    history_url = f"https://fantasy.premierleague.com/api/element-summary/{player_id}/"
    history_response = requests.get(history_url)
    history_data = history_response.json()
    
    # Get the last 4 games
    last_4_games = history_data['history'][-4:]

    # Initialize stats for the player
    total_xg = total_xa = total_xgc = total_points = total_minutes = total_bonus = total_saves = 0

    # Loop through the last 4 games and accumulate the stats
    for game in last_4_games:
        total_xg += float(game.get('expected_goals', 0))
        total_xa += float(game.get('expected_assists', 0))
        total_xgc += float(game.get('expected_goals_conceded', 0))
        total_points += game.get('total_points', 0)
        total_minutes += game.get('minutes', 0)
        total_bonus += game.get('bonus', 0)
        if position == 1:  # Goalkeeper
            total_saves += game.get('saves', 0)

    # Calculate averages for the stats
    avg_xg = total_xg / 4
    avg_xa = total_xa / 4
    avg_xgc = total_xgc / 4
    avg_points = total_points / 4
    avg_minutes = total_minutes / 4
    avg_bonus = total_bonus / 4
    avg_saves = total_saves / 4 if position == 1 else 0
    minutes_category = 0 if avg_minutes == 0 else (1 if 0 < avg_minutes < 60 else 2)
    p_x0 = poisson_prob(avg_xgc, 0)

    # Get pFDR and fFDR
    pFDR, fFDR = calculate_fdr(team_id, fixtures_data)

    # Calculate xValue and Value for player
    if position == 3:  # Midfielder
        cs_points = p_x0 if minutes_category == 2 else 0
        xppg = 5 * avg_xg + 3 * avg_xa + minutes_category + avg_bonus + cs_points
        xValue = xppg / (player['now_cost'] / 10)
        if xValue > 0:  # Only add players with xValue > 0
            player_data["Midfielders"].append([full_name, avg_xg, avg_xa, avg_xgc, avg_bonus, avg_minutes, xppg, avg_points, player['now_cost'] / 10, avg_points / (player['now_cost'] / 10), xValue, pFDR, fFDR])
    elif position == 2:  # Defender
        xppg = 6 * avg_xg + 3 * avg_xa + minutes_category + avg_bonus
        if avg_minutes >= 60:
            xppg += 4 * p_x0  # Add clean sheet points if minutes >= 60
            for i in range(2, 15, 2):
                xppg -= poisson_prob(avg_xgc, i)
        xValue = xppg / (player['now_cost'] / 10)
        if xValue > 0:  # Only add players with xValue > 0
            player_data["Defenders"].append([full_name, avg_xg, avg_xa, avg_xgc, avg_bonus, avg_minutes, xppg, avg_points, player['now_cost'] / 10, avg_points / (player['now_cost'] / 10), xValue, pFDR, fFDR])
    elif position == 4:  # Attacker
        xppg = 4 * avg_xg + 3 * avg_xa + minutes_category + avg_bonus
        xValue = xppg / (player['now_cost'] / 10)
        if xValue > 0:  # Only add players with xValue > 0
            player_data["Attackers"].append([full_name, avg_xg, avg_xa, avg_bonus, avg_minutes, xppg, avg_points, player['now_cost'] / 10, avg_points / (player['now_cost'] / 10), xValue, pFDR, fFDR])
    elif position == 1:  # Goalkeeper
        xppg = 3 * avg_xa + minutes_category + avg_bonus + (avg_saves / 3)
        if avg_minutes >= 60:
            xppg += 4 * p_x0  # Add clean sheet points if minutes >= 60
            for i in range(2, 15, 2):
                xppg -= poisson_prob(avg_xgc, i)
        xValue = xppg / (player['now_cost'] / 10)
        if xValue > 0:  # Only add players with xValue > 0
            player_data["Goalkeepers"].append([full_name, avg_xa, avg_xgc, avg_bonus, avg_minutes, avg_saves, xppg, avg_points, player['now_cost'] / 10, avg_points / (player['now_cost'] / 10), xValue, pFDR, fFDR])

# Create separate sheets for each category
categories = ["Goalkeepers", "Defenders", "Attackers", "Midfielders"]

# Headers for the stats (with Value before xValue)
for category in categories:
    # Create or get the sheet for this category
    if category in workbook.sheetnames:
        sheet = workbook[category]
    else:
        sheet = workbook.create_sheet(title=category)

    # Clear the sheet in case it already exists
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None

    # Write the headers in the second row
    if category == "Attackers":
        headers = ["Rank", "Player Name", "xG", "xA", "BP", "Minutes", "xPPG", "Points", "Price (£)", "Value", "xValue", "pFDR", "fFDR"]
    elif category == "Goalkeepers":
        headers = ["Rank", "Player Name", "xA", "xGC", "BP", "Minutes", "Avg Saves", "xPPG", "Points", "Price (£)", "Value", "xValue", "pFDR", "fFDR"]
    elif category == "Midfielders":
        headers = ["Rank", "Player Name", "xG", "xA", "xGC", "BP", "Minutes", "xPPG", "Points", "Price (£)", "Value", "xValue", "pFDR", "fFDR"]
    elif category == "Defenders":
        headers = ["Rank", "Player Name", "xG", "xA", "xGC", "BP", "Minutes", "xPPG", "Points", "Price (£)", "Value", "xValue", "pFDR", "fFDR"]

    for col_num, header in enumerate(headers, start=2):  # Start from column B
        sheet.cell(row=2, column=col_num, value=header)

    # Apply bold formatting to headers
    for col_num in range(2, len(headers) + 2):
        sheet.cell(row=2, column=col_num).font = Font(bold=True)

    # Sort players by xValue in descending order and write player data for this category
    player_data[category].sort(key=lambda x: x[-3], reverse=True)  # Sort by xValue (third last element)

    row = 3  # Start writing data from the third row
    for rank, player in enumerate(player_data[category], start=1):
        sheet.cell(row=row, column=2, value=rank)  # Rank in column B
        for col_num, value in enumerate(player, start=3):  # Data starts from column C
            sheet.cell(row=row, column=col_num, value=value)
        row += 1

# Save the updated workbook
workbook.save("players_data.xlsx")
print("Players' data saved to separate sheets in 'players_data.xlsx'")
